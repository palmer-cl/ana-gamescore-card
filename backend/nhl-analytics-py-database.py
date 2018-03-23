import bs4 as bs
import urllib
import contextlib
import os
from openpyxl import load_workbook
import pymysql.cursors

# ##########################################################################################################################
#
#                                Database connection
#
# ##########################################################################################################################

conn = pymysql.connect(
    host='localhost',
    user='root',
    password='',
    db='qsio_db',
    cursorclass=pymysql.cursors.DictCursor)


# ##########################################################################################################################
#
#                                Part 1A: Individual Game Data
#
# ##########################################################################################################################
print os.getcwd()
##Game number in URL##
GameNum='21135'

##Build full URL##
GAMELINK = ('http://www.naturalstattrick.com/game.php?season=20172018&game=' + GameNum + '&view=limited')

##Open the Excel Workbook##
wb = load_workbook(filename="nhl-test-py.xlsx")
sht1= wb['AllSit-Individual']
sht2= wb['5v5-On-Ice']
sht3= wb['AllSit- Goalies']
sht4= wb['Season GameScore']

##Build 4 arrays ##
game_info=[]
master_on_ice_5v5=[]
master_ind_all_sit=[]
master_G_ind_all_sit=[]


with contextlib.closing(urllib.urlopen(GAMELINK)) as source:
    soup = bs.BeautifulSoup(source, "lxml")

##Get game date and opponent and print to the screen##
for i in soup.find_all('h1'):
    h1text = i.text
VisitingTeam, HomeTeam = h1text.split(" @ ")

##Get game time and score##
for i in soup.find_all('h2'):
    h2gamedate = i.text

head, sep, tail = h2gamedate.partition('\n')
h2gamedate = head

print tail

tail.encode("utf-8")
a,b,c = tail.split(" ")
c=c.split("\n")

VisitingScore = a
HomeScore = c[0]

game_info.append(h2gamedate.encode("utf-8"))
game_info.append(VisitingTeam.encode("utf-8"))
game_info.append(HomeTeam.encode("utf-8"))
game_info.append(VisitingScore.encode("utf-8"))
game_info.append(HomeScore.encode("utf-8"))
print game_info

#Save the Game Information to SQL##

with conn.cursor() as cursor:
    # Create a new record
    placeholders = ', '.join(['%s'] * len(game_info))  # "%s, %s, %s, ... %s"
    query = 'INSERT INTO `game_info` (`GameDate`, `vis_team`, `home_team`, `vis_score`, `home_score`) VALUES ({})'.format(
        placeholders)
    cursor.execute(query, tuple(game_info))

# connection is not autocommit by default. So you must commit to save
# your changes.
conn.commit()


##Prompt screen when first data set starts##
print "Starting Player gather data for first set:"
#Define a function for gathering only player data (goalies seperate#
def grabSkaterData(className, idName, list):
    for master in soup.find_all("div", {"class": className}):
        for i in master.find_all("table", {"id": idName}):

            for record in i.find_all('tr'):
                playerdata = []
                for data in record.find_all('td'):
                    new = (data.text).encode('utf-8')
                    new = new.replace('\xa0', ' ')
                    new = new.replace('\xc2', '')
                    playerdata.extend([new])

                # Print to the screen each players data gathered##
                print playerdata

                ##Manipulate the data to add the date in the first column##
                playerdata.insert(0, str(h2gamedate))
                list.append(playerdata)
        list.pop(0)
        break

    print list

    return

##Print to the screen gathering data from all situations##
print "Printing ALL SITUATIONS from div class and table id: "
grabSkaterData("tall", "tbANAstall", master_ind_all_sit)
print '\n'


##Print to the screen gathering data from only on ice situations##
print "Printing 5v5 ON ICE from div class and table id: "
grabSkaterData("t5v5", "tbANAoi5v5", master_on_ice_5v5)
print '\n'

##Print to the screen gathering data from the goalie class##
print '\n'
print "Printing ALL SITUATIONS- GOALIES from div class and table id: "

for master in soup.find_all("div", {"class": "tall"}):
    for i in master.find_all("table", {"id": "tbANAstgall"}):

        goalieCounter=0

        for record in i.find_all('tr'):
            playerdata = []
            for data in record.find_all('td'):
                new = (data.text).encode('utf-8')
                new = new.replace('\xa0', ' ')
                new = new.replace('\xc2', '')
                playerdata.extend([new])

            master_G_ind_all_sit.append(playerdata)


##Remove empty lists from the list##
master_G_ind_all_sit = [x for x in master_G_ind_all_sit if x!= []]

## 2 Goalie Possible list##
##Declare list##
TwoGoaliePossibleList=[]

Goalie1= master_G_ind_all_sit[0]
Goalie2=master_G_ind_all_sit[1]

print Goalie1
print Goalie2

##For Goalie 1 ##
##delete unwanted information and keep only the stats we want##
del Goalie1[6]
del Goalie1[5]
del Goalie1[3]
del Goalie1[1]
##Insert Game Date and Position
Goalie1.insert(0,str(h2gamedate))
Goalie1.insert(2,str('G'))

##For Goalie 2 ##
##delete unwanted information and keep only the stats we want##
del Goalie2[6]
del Goalie2[5]
del Goalie2[3]
del Goalie2[1]
##Insert Game Date and Position
Goalie2.insert(0,str(h2gamedate))
Goalie2.insert(2,str('G'))

print Goalie1
print Goalie2

FinalGoalieList= []

## Check if one goalie played or two goalies played ##
if Goalie1[1] == Goalie2[1]:
    FinalGoalieList=Goalie1
    print "ONE GOALIE PLAYED THIS GAME."
else:
    FinalGoalieList.append(Goalie1)
    FinalGoalieList.append(Goalie2)
    print "2 GOALIES PLAYED IN THIS GAME"

print FinalGoalieList

#APPEND PROPER HEADERS FOR THE LISTS ##
IND_header = ['GameDate','Player', 'Position', 'TOI', 'Goals', 'Total Assists', 'First Assists', 'Second_Assists', 'Total Points', 'Shots', 'SH%', 'iCF', 'iSCF', 'iHDCF', 'Rebounds Created', 'PIM' , 'Total Penalties', 'Minors', 'Majors', 'Misconduct', 'Pen Drawn', 'Giveaways', 'Takeaways', 'Hits', 'Hits Taken', 'Shots Blocked', 'Faceoffs Won', 'Faceoffs Lost', 'Faceoff %']
OI_header = ['GameDate','Player', 'Position', 'TOI', 'CF', 'CA', 'CF%', 'CF% Rel', 'FF', 'FA', 'FF%', 'FF% Rel', 'SF', 'SA', 'SF%', 'SF% Rel', 'GF', 'GA', 'GF%', 'GF% Rel', 'SCF', 'SCA', 'SCF%', 'SCF% Rel', 'HDCF', 'HDCA', 'HDCF%', 'HDCF Rel', 'Off. Zone Faceoffs', 'Neu. Zone Faceoffs', 'Def. Zone Faceoffs', 'Off Zone Faceoff %']
G_header = ['GameDate', 'Player', 'Position', 'Goals', 'Shots']

## Leave this in case we need to redo the header...##
#master_ind_all_sit.insert(0,IND_header)
#master_on_ice_5v5.insert(0,OI_header)
#G_list_combined.insert(0,G_header)
#G_list_combined.insert(1,master_g_only_list)

def replace(l):
    for x in l:
      for i,v in enumerate(x):
         if v == '-':
            x.pop(i)
            x.insert(i, '0')

replace(master_on_ice_5v5)
replace(master_ind_all_sit)

# Make sure the lists have no - values ##
print '\n'
print "Here are the new build lists with proper values:"
print master_ind_all_sit
print master_on_ice_5v5
print FinalGoalieList
print '\n'


##Goalies First##
##Special circumstances, because goalies##
if len(FinalGoalieList) == 2:
    for i in FinalGoalieList:
        with conn.cursor() as cursor:
            # Create a new record
            placeholders = ', '.join(['%s'] * len(i))  # "%s, %s, %s, ... %s"
            query = 'INSERT INTO `allsit_goalies` (`GameDate`, `Player`, `Position`, `shots`, `goals`) VALUES ({})'.format(
                placeholders)
            cursor.execute(query, tuple(i))
        conn.commit()
else:
    with conn.cursor() as cursor:
        # Create a new record
        placeholders = ', '.join(['%s'] * len(FinalGoalieList))  # "%s, %s, %s, ... %s"
        query = 'INSERT INTO `allsit_goalies` (`GameDate`, `Player`, `Position`, `shots`, `goals`) VALUES ({})'.format(
            placeholders)
        cursor.execute(query, tuple(FinalGoalieList))
    conn.commit()

# ##On Ice list next##
for i in master_on_ice_5v5:
    with conn.cursor() as cursor:
        # Create a new record
        placeholders = ', '.join(['%s'] * len(i))  # "%s, %s, %s, ... %s"
        query = 'INSERT INTO `gs_5v5_oi` (`GameDate`, `Player`, `Position`, `TOI`, `CF`, `CA`, `CFpercent`, `CFpercent_Rel`,' \
                ' `FF`, `FA`, `FFpercent`, `FFpercent_Rel`, `SF`, `SA`, `SFpercent`, `SFpercent_Rel`, `GF`, `GA`, `GFpercent`,' \
                ' `GFpercent_Rel`, `SCF`, `SCA`, `SCFpercent`, `SCFpercent_Rel`, `HDCF`, `HDCA`, `HDCFpercent`, `HDCF_Rel`, `OZ_Faceoffs`,' \
                ' `NU_Faceoffs`, `DZ_Faceoffs`, `OZ_Faceoffpercent`) VALUES ({})'.format(
            placeholders)
        cursor.execute(query, tuple(i))
    conn.commit()

for i in master_ind_all_sit:
    with conn.cursor() as cursor:
        # Create a new record
        placeholders = ', '.join(['%s'] * len(i))  # "%s, %s, %s, ... %s"
        query = 'INSERT INTO `gs_allsit_ind` (`GameDate`, `Player`, `Position`, `TOI`, `Goals`, `Total_Assists`, ' \
                '`First_Assists`, `Second_Assists`, `Total_Points`, `Shots`, `SHpercent`, `iCF`, `iSCF`, `iHDCF`, ' \
                '`Rebounds_Created`, `PIM`, `Total_Penalties`, `Minors`, `Majors`, `Misconduct`, `Pen_Drawn`, `Giveaways`,' \
                ' `Takeaways`, `Hits`, `Hits_Taken`, `Shots_Blocked`, `Faceoffs_Won`, `Faceoffs_Lost`, `Faceoffpercent`) ' \
                'VALUES ({})'.format(
            placeholders)
        cursor.execute(query, tuple(i))
    conn.commit()








# # ##########################################################################################################################
# #
# #                                Part 1B: Start Printing to Excel
# #
# # ##########################################################################################################################
#
# ##Define a function for writing to Excel Rows ##
# def writeToRows(sheet,list):
#     row = 1
#     col = 1
#     while sheet.cell(row=row, column=1).value != None:
#         row += 1
#
#     for i in list:
#         col = 1
#         for element in i:
#             if col > 3:
#                 if element == '-':
#                     element = 0
#                 sheet.cell(row=row, column=col).value = float(element)
#                 col += 1
#             else:
#                 if element == '-':
#                     element = 0
#                 sheet.cell(row=row, column=col).value = element
#                 col += 1
#         row += 1
#
#     return
#
#
# ##COPY TO EXCEL##
# print "Part 1: Starting to copy to Excel now..."
# writeToRows(sht1,master_ind_all_sit)
# print "Part 1: All Situations EXCEL complete..."
#
# print "Part 2: Starting to copy to Excel now..."
# writeToRows(sht2,master_on_ice_5v5)
# print "Part 2: On Ice Data EXCEL complete..."
#
# print "Part 3: Starting to copy to Excel now..."
#
# row = 1
# col = 1
# while (sht3.cell(row=row,column=1).value != None):
#     row+=1
#
# if len(FinalGoalieList) == 5:
#     for i in FinalGoalieList:
#         if col > 3:
#             if i == '-':
#                 i = 0
#             sht3.cell(row=row,column=col).value = float(i)
#             col +=1
#         else:
#             if i == '-':
#                 i = 0
#             sht3.cell(row=row, column=col).value = i
#             col += 1
# else:
#     for i in FinalGoalieList:
#         col=1
#         for e in i:
#             if col >3:
#                 if i == '-':
#                     i = 0
#                 sht3.cell(row=row, column=col).value = float(e)
#                 col +=1
#             else:
#                 if i == '-':
#                     i = 0
#                 sht3.cell(row=row, column=col).value = e
#                 col +=1
#
#         row+=1
#
#     print "Part 3: GOALIES to EXCEL complete..."
#
# ### Print final part of season game score sheets ###
# # ##########################################################################################################################
# #
# #                                Part 2 : Season Data
# #
# # ##########################################################################################################################
# # # ##COPY TO EXCEL##
# master_on_ice=[]
# master_ind=[]
# header1= ['Player', 'Position', 'GP', '5v5oi_TOI', '5v5oi_CF', '5v5oi_CA', '5v5oi_CF%', '5v5oi_FF', '5v5oi_FA', '5v5oi_FF%',
#          '5v5oi_SF', '5v5oi_SA', '5v5oi_SF%', '5v5oi_GF', '5v5oi_GA', '5v5oi_GF%', '5v5oi_SCF', '5v5oi_SCA', '5v5oi_SCF%',
#           '5v5oi_HDCF', '5v5oi_HDCA', '5v5oi_HDCF%', '5v5oi_HDGF', '5v5oi_HDGA', '5v5oi_HDGF%', '5v5oi_MDCF', '5v5oi_MDCA',
#           '5v5oi_MDCF%', '5v5oi_MDGF', '5v5oi_MDGA', '5v5oi_MDGF%', '5v5oi_LDCF', '5v5oi_LDCA', '5v5oi_LDCF%', '5v5oi_LDGF',
#           '5v5oi_LDGA', '5v5oi_LDGF%', '5v5oi_On-Ice SH%', '5v5oi_On-Ice SV%', '5v5oi_PDO', '5v5oi_Off. Zone Faceoffs',
#           '5v5oi_Neu. Zone Faceoffs', '5v5oi_Def. Zone Faceoffs', '5v5oi_Off Zone Faceoff %']
# header2 = ['Player', 'Position', 'GP', 'TOI', 'Goals', 'Total Assists', 'First Assists', 'Sencond Assists', 'Total Points',
#            'Shots', 'SH%', 'iCF', 'iSCF', 'iHDCF', 'Rebounds Created', 'PIM' , 'Total Penalities', 'Minors', 'Majors',
#            'Misconduct', 'Pen Drawn', 'Giveaways', 'Takeaways', 'Hits', 'Hits Taken', 'Shots Blocked','Faceoffs Won',
#            'Faceoffs Lost', 'Faceoff %']
#
# def resetSeasonData():
#     playerdata = playerdatasaved = ""
#     counter = 1
#
#     return
#
# def grabSeasonData(head,list):
#     counter = 1
#
#     for record in soup.find_all('tr'):
#         playerdata = []
#         for data in record.find_all('td'):
#             new = (data.text).encode('utf-8')
#             new = new.replace('\xa0', ' ')
#             new = new.replace('\xc2', '')
#             playerdata.extend([new])
#
#         if counter == 1:
#             list.append(head)
#
#         else:
#             playerdata.pop(0)
#             list.append(playerdata)
#         counter += 1
#
#     return
#
# ##Second set of team data (On-Ice) ###
# print "Gather Season data starting now..."
# with contextlib.closing(urllib.urlopen('https://www.naturalstattrick.com/playerteams.php?season=20172018&stype=2&sit=5v5&score=all&stdoi=oi&rate=n&team=ANA&pos=S&loc=B&toi=0&gpfilt=none&fd=&td=&tgp=82&lines=single')) as source:
#     soup = bs.BeautifulSoup(source, "html.parser")
#
# ## Reset and gather new data ##
# resetSeasonData()
# grabSeasonData(header1,master_on_ice)
#
#
# ##Second set of team data (Individual) ###
# source = urllib.urlopen('https://www.naturalstattrick.com/playerteams.php?season=20172018&stype=2&sit=all&score=all&stdoi=std&rate=n&team=ANA&pos=S&loc=B&toi=0&gpfilt=none&fd=&td=&tgp=82&lines=single').read()
# soup = bs.BeautifulSoup(source, "html.parser")
#
# ## Reset and gather new data ##
# resetSeasonData()
# grabSeasonData(header2,master_ind)
#
#
# ## Combining the list using the map function with lambda ##
# #Confirm our list was combined into one long list
# rawcombined = map(lambda a, b: a + b, master_on_ice, master_ind)
#
# #Remove the data we don't want to keep#
# for i in rawcombined:
#     del i[44:47]
#
# ##Print season to excel sheet before google sheet ###
# row = 1
# for element in rawcombined:
#     col = 1
#     if row==1:
#         print "Skip Header"
#     else:
#         for i in element:
#             if col > 2:
#                 sht4.cell(row=row,column=col).value = float(i)
#                 col += 1
#             else:
#                 sht4.cell(row=row, column=col).value = i
#                 col += 1
#
#     row +=1
#
# wb.save('nhl-test-py.xlsx')
conn.close()
print "All processes complete."
print "*****DB SAVED****"
